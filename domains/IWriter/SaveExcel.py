from domains.my_additional import *
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import get_column_letter

class SaveExcel:

    @classmethod
    def save(cls, name_domain,  response):
        try:
            if not os.path.exists(PATH_EXCEL_FILE):
                wb = Workbook()
            else:
                wb = load_workbook(PATH_EXCEL_FILE)
            ws = wb.active

            results = cls.formed_data(response)

            column = ws.max_column+1
            for row, value in enumerate(results, 1):
                if row == 12:
                    cls.add_screenshot_excel(ws, row, column, name_domain)
                else:
                    ws.cell(row=row, column=column, value=str(value))

            wb.save(PATH_EXCEL_FILE)
        except Exception as err:
            logger.error(f"Ошибка при обработке данных для Excel: {str(err)}")




    @staticmethod
    def formed_header():
        return (
            "Анализируемый домен", "Контакты", "IP-адрес", "Владелец", "Доступность сайта", "Регистратор", "Создан", "На последнем ip-адресе",
            "Оплачен до", "Статус","Список DNS-серверов", "Скриншот", "Наличие архивных версий сайта", 'Связанные адреса эл.почты', 'Поддомены',
            'Данные указанные на сайте', 'IP-адреса на которых размещался домен', 'Прочие домены на сервере',
            'Прочие домены принадлежащие владельцу'
        )


    @staticmethod
    def formed_data(response):
        try:
            if "main" in response:
                main = response["main"]
                results = [
                    main.get("domain",""),
                    main.get("ip",""),
                    main.get("owner",""),
                    response.get('state'),
                    main.get("registrar",""),
                    main.get("create_time",""),
                    main.get("on_last_ip_address",""),
                    main.get("expires",""),
                    main.get("status",""),
                    field_ns_server(main.get("ns","")),
                    "Отсутствует",
                    main.get("archive", ""),
                    linked_email_save(main.get("linked_email", "")),
                    worker_write_list(main.get("subdomains")),
                    dict_to_str(main.get("metadata")),
                           ]

                result = worker_write_list(main.get("contact"))
                results.insert(1,result)

                result = worker_dynamic_type(main.get("ip_history"))
                results.append(result)

                result = worker_dynamic_type(main.get("other_domain"), typeList=True)
                results.append(result)

                result = worker_dynamic_type(main.get("rw_domain"), typeList=True)
                results.append(result)

                return results

        except Exception as err:
            logger.error(f"Ошибка при запуске обработчика Excel: {str(err)}")


    @classmethod
    def finally_xls(cls):
        wb = load_workbook(PATH_EXCEL_FILE)
        ws = wb.active

        header = cls.formed_header()

        for row, value in enumerate(header, 1):
            ws.cell(row=row, column=1, value=value)
        ws.freeze_panes = "A1"

        ft = Font(bold=True)
        al = Alignment(wrap_text=True, vertical='top')


        for i in range(1, ws.max_column + 1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = 50

        for cols in ws.rows:
            for num, cell in enumerate(cols, 1):
                if num == 1:
                    cell.font = ft
                cell.alignment = al

        wb.save(PATH_EXCEL_FILE)

    @staticmethod
    def add_screenshot_excel(ws, row, column, name_domain):
        try:
            PATH_SCREENSHOT = os.path.join(PATH_RESULTS_DIR, name_domain, "screenshot.png")
            if os.path.exists(PATH_SCREENSHOT):
                screenshot = Image(PATH_SCREENSHOT)
                screenshot.height = 200
                screenshot.width = 370
                column = get_column_letter(column)
                ws.row_dimensions[row].height = 150
                ws.add_image(screenshot, column + str(row))
        except:
            logger.error("Ошибка при вставке скриншота для Excel")



