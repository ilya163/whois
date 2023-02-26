from .my_additional import *
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

class saveExcel:
    def __init__(self, domain):
        self.domain = domain

    def run(self, response):
            try:
                result = {}
                for name in response:
                    for block in response[name]:
                        if block in result and type(result[block]) == dict:
                            result[block].update(response[name][block])
                        elif block in result and type(result[block]) == list:
                            result[block].append(response[name][block][0])
                        else:
                            result[block] = response[name][block]

                self.save_xls(result)
                # logger.info("Данные в формате Excel записаны")
            except Exception as err:
                logger.error(f"Ошибка при записи в Excel файл: {str(err)}")



    def xls_worker(self, ws, response):
        if "main" in response:
            main = response["main"]
            ws["A1"], ws["B1"] = "Анализируемый домен", main["domain"]
            ws["A2"], ws["B2"] = "IP-адрес", main["ip"]
            ws["A3"], ws["B3"] = "Владелец", main["owner"]
            ws["A4"], ws["B4"] = "Регистратор", main["registrar"]
            ws["A5"], ws["B5"] = "Создан", main["create_time"]
            ws["A6"], ws["B6"] = "На последнем ip-адресе", main["on_last_ip_address"]
            ws["A7"], ws["B7"] = "Оплачен до", main["expires"]
            ws["A8"], ws["B8"] = "Контакты", main["contact"]
            ws["A9"], ws["B9"] = "Статус", main["status"]
            ws["A10"], ws["B10"] = "Список DNS-серверов", field_ns_server(main["ns"])

            if "ip_history" in main:
                if len(main["ip_history"]) > 0:
                    result = xls_worker_dynamic_type(main["ip_history"],
                                                                   header='IP-адреса на которых размещался домен')
                    for xls in result:
                        ws.append(xls)

            if "other_domain" in main:
                if len(main["other_domain"]) > 0:
                    result = xls_worker_dynamic_type(main["other_domain"],
                                                                   header='Прочие домены на сервере')
                    ws.append(result)


            if "rw_domain" in main:
                if len(main["rw_domain"]) > 0:
                    result = xls_worker_dynamic_type(main["rw_domain"],
                                                                   header='Прочие домены принадлежащие владельцу')
                    for xls in result:
                        ws.append(xls)
                ws.append([])
                ws.append([])

        results = {}
        if "ip_history_all" in response:
            if len(response["ip_history_all"]) > 0:
                results["ip_history_all"] = xls_worker_dynamic_type_all(response["ip_history_all"],
                                                                                      header=[
                                                                                          'IP-адреса на которых размещался домен',
                                                                                          'Локация', 'Владелец',
                                                                                          'Последние изменения'])
                for xls in results["ip_history_all"]:
                    ws.append(xls)
                ws.append([])
                ws.append([])

        if "rw_domain_all" in response:
            if len(response["rw_domain_all"]) > 0:
                results["rw_domain_all"] = xls_worker_dynamic_type_all(response["rw_domain_all"],
                                                                                     header=[
                                                                                         'Прочие домены принадлежащие владельцу',
                                                                                         'Дата создания',
                                                                                         'Регистратор'])

                for xls in results["rw_domain_all"]:
                    ws.append(xls)
                ws.append([])
                ws.append([])

        if "other_domain_all" in response:
            if len(response["other_domain_all"]) > 0:
                results["other_domain_all"] = xls_worker_dynamic_type_all(response["other_domain_all"],
                                                                                        header=[
                                                                                            "Прочие домены на сервере"])
                for xls in results["other_domain_all"]:
                    ws.append(xls)
                ws.append([])
                ws.append([])
        return ws

    def save_xls(self, response):
            try:
                if not os.path.exists(PATH_EXCEL_FILE):
                    wb = openpyxl.Workbook()
                    wb.save(PATH_EXCEL_FILE)
                wb = load_workbook(PATH_EXCEL_FILE)
                if "Sheet" in wb:
                    del wb["Sheet"]
                ws = wb.create_sheet(self.domain)
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 50

                self.xls_worker(ws, response)

                ft = Font(bold=True)
                al = Alignment(wrap_text=True)
                for row in ws["A1:A20"]:
                    for cell in row:
                        cell.font = ft
                for row in ws["B1:B50"]:
                    for cell in row:
                        cell.alignment = al

                wb.save(PATH_EXCEL_FILE)
                # logger.info("Данные в формате Excel записаны")
            except Exception as err:
                logger.error(f"Ошибка при записи в Excel файл: {str(err)}")